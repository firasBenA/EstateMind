"""
STEP 1 - IPIM Data Loader
===========================
Lit directement le fichier Excel INS placé dans le même dossier que ce script.
Aucune connexion internet. Aucune saisie manuelle.

Usage:
  1. Placer ce script et le fichier Excel dans le même dossier
  2. python step1_collect_ipim.py
  3. Le CSV résultant est créé dans le même dossier

Fichier Excel attendu (nom flexible, doit contenir "IPIM" ou "Series"):
  Ex: Series_d_indices_IPIM_de_T1-2000_à_T1-2024.xlsx
"""

import sys
import glob
from pathlib import Path
import pandas as pd
import openpyxl

# ── Tout se passe dans le même dossier que ce script ─────────────────────────
SCRIPT_DIR = Path(__file__).parent
OUTPUT_CSV = SCRIPT_DIR / "ipim_historical.csv"

SHEET_INDICES = "Indices par type à diffuser "
SHEET_VOLUME  = "Gli. Effec. par type à diffuser"


def find_excel():
    """Cherche le fichier Excel IPIM dans le même dossier que ce script."""
    patterns = ["*IPIM*.xlsx", "*ipim*.xlsx", "*Series*indices*.xlsx", "*Series*IPIM*.xlsx"]
    for pattern in patterns:
        matches = list(SCRIPT_DIR.glob(pattern))
        if matches:
            return matches[0]
    return None


def parse_sheet(ws, data_start_row, col_indices):
    """
    Parse générique : lit les lignes à partir de data_start_row,
    gère les années fusionnées (None = même année que la ligne précédente).
    col_indices = dict {nom: index_colonne}
    """
    rows = list(ws.iter_rows(values_only=True))
    records = []
    current_year = None

    for row in rows[data_start_row:]:
        year_val = row[col_indices["year"]]
        qtr_val  = row[col_indices["quarter"]]

        if year_val is not None:
            try:
                current_year = int(year_val)
            except (TypeError, ValueError):
                continue

        if qtr_val is None or current_year is None:
            continue

        try:
            q = int(qtr_val)
        except (TypeError, ValueError):
            continue

        record = {"year": current_year, "quarter": q}
        for name, idx in col_indices.items():
            if name in ("year", "quarter"):
                continue
            val = row[idx] if idx < len(row) else None
            record[name] = float(val) if val is not None else None

        # Ignorer les lignes entièrement vides (hors year/quarter)
        data_cols = [k for k in record if k not in ("year", "quarter")]
        if all(record[k] is None for k in data_cols):
            continue

        records.append(record)

    return pd.DataFrame(records)


def find_header_row(ws, keyword="trimestre"):
    """Trouve l'index de la ligne d'en-tête contenant un mot-clé."""
    rows = list(ws.iter_rows(max_row=15, values_only=True))
    for i, row in enumerate(rows):
        row_str = " ".join(str(c or "").lower() for c in row)
        if keyword in row_str and "ann" in row_str:
            return i
    return 7  # fallback


if __name__ == "__main__":
    print("=" * 55)
    print("STEP 1 — Chargement IPIM depuis Excel INS")
    print("=" * 55)

    # ── Trouver le fichier Excel ──────────────────────────────
    xl_path = find_excel()
    if xl_path is None:
        print(f"\n❌ Aucun fichier Excel IPIM trouvé dans:\n   {SCRIPT_DIR}")
        print("\n   Placer le fichier Excel INS dans ce dossier.")
        print("   Le nom doit contenir 'IPIM' ou 'Series'.")
        sys.exit(1)

    print(f"\n📂 Fichier: {xl_path.name}")
    wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
    print(f"   Feuilles: {wb.sheetnames}")

    # ── Feuille 1 : Niveaux d'indices ────────────────────────
    print(f"\n📊 Lecture feuille 1 — Indices (base 2015=100)...")
    ws1 = wb[SHEET_INDICES]
    header_row = find_header_row(ws1)

    df_idx = parse_sheet(ws1, data_start_row=header_row + 1, col_indices={
        "year":             0,
        "quarter":          1,
        "ipim_terrain":     2,
        "ipim_appartement": 3,
        "ipim_maison":      4,
    })

    df_idx["date"]   = df_idx.apply(lambda r: pd.Timestamp(year=int(r.year), month=(int(r.quarter)-1)*3+1, day=1), axis=1)
    df_idx["period"] = df_idx.apply(lambda r: f"{int(r.year)}-Q{int(r.quarter)}", axis=1)

    # Arrondir les indices
    for col in ["ipim_terrain", "ipim_appartement", "ipim_maison"]:
        df_idx[col] = df_idx[col].round(4)

    # Bâti total = pondération officielle ~75% appart / 25% maison
    df_idx["ipim_bati_total"] = (0.75 * df_idx["ipim_appartement"] + 0.25 * df_idx["ipim_maison"]).round(4)

    print(f"   → {len(df_idx)} trimestres ({df_idx['period'].iloc[0]} → {df_idx['period'].iloc[-1]})")

    # ── Feuille 2 : Volume transactions CVS ──────────────────
    print(f"\n📊 Lecture feuille 2 — Volume transactions CVS (%)...")
    ws2 = wb[SHEET_VOLUME]
    header_row2 = find_header_row(ws2)

    df_vol = parse_sheet(ws2, data_start_row=header_row2 + 1, col_indices={
        "year":                 0,
        "quarter":              1,
        "vol_terrain_pct":      2,
        "vol_appartement_pct":  3,
        "vol_maison_pct":       4,
    })
    df_vol["period"] = df_vol.apply(lambda r: f"{int(r.year)}-Q{int(r.quarter)}", axis=1)

    # Les valeurs sont en décimal (0.10 = 10%) → convertir en %
    for col in ["vol_terrain_pct", "vol_appartement_pct", "vol_maison_pct"]:
        df_vol[col] = (df_vol[col] * 100).round(4)

    print(f"   → {len(df_vol)} trimestres")

    # ── Merge et export ──────────────────────────────────────
    df_final = df_idx.merge(
        df_vol[["period", "vol_terrain_pct", "vol_appartement_pct", "vol_maison_pct"]],
        on="period", how="left"
    ).sort_values("date").reset_index(drop=True)

    # Colonnes finales ordonnées
    cols = ["date", "year", "quarter", "period",
            "ipim_terrain", "ipim_appartement", "ipim_maison", "ipim_bati_total",
            "vol_terrain_pct", "vol_appartement_pct", "vol_maison_pct"]
    df_final = df_final[cols]

    df_final.to_csv(OUTPUT_CSV, index=False)

    # ── Résumé ───────────────────────────────────────────────
    print(f"\n✅ CSV généré: {OUTPUT_CSV.name}")
    print(f"   {len(df_final)} trimestres | {df_final['period'].iloc[0]} → {df_final['period'].iloc[-1]}")
    print(f"\nDerniers trimestres:")
    display_cols = ["period", "ipim_terrain", "ipim_appartement", "ipim_maison", "ipim_bati_total"]
    print(df_final[display_cols].tail(8).to_string(index=False))